import os
import pandas as pd
from openpyxl import load_workbook

# Ruta de la carpeta con los archivos
ruta_carpeta = r"C:\Users\rodri\Downloads\compilacion"

# Lista para almacenar los datos combinados
datos_combinados = []

# Columnas a extraer (basadas en índices de Excel: A=0, B=1, ..., K=10)
columnas_a_extraer = list(range(0, 11))  # Índices de las columnas A-K

# Recorre todos los archivos en la carpeta
for archivo in os.listdir(ruta_carpeta):
    ruta_archivo = os.path.join(ruta_carpeta, archivo)
    try:
        if archivo.endswith(".xlsx"):
            # Procesar archivos Excel
            wb = load_workbook(ruta_archivo, data_only=True)
            for hoja in wb.sheetnames:
                if hoja == "MLM-CARS_AND_VANS":  # Saltar la hoja especificada
                    continue

                ws = wb[hoja]
                for fila in ws.iter_rows(min_row=2, max_col=11, values_only=True):  # Saltar encabezados
                    if any(fila):  # Verificar si hay datos en la fila
                        datos_combinados.append(fila)

        elif archivo.endswith(".csv"):
            # Procesar archivos CSV
            df = pd.read_csv(ruta_archivo, usecols=range(0, 11))  # Leer columnas A-K
            datos_combinados.extend(df.values.tolist())

    except Exception as e:
        print(f"Error al procesar el archivo {archivo}: {e}")

# Crear un DataFrame con los datos combinados
df_combinado = pd.DataFrame(datos_combinados, columns=["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"])

# Dividir el DataFrame en partes de 500,000 filas y guardar en un archivo Excel
ruta_salida = os.path.join(ruta_carpeta, "combinado.xlsx")
with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
    for i in range(0, len(df_combinado), 500000):
        df_combinado.iloc[i:i+500000].to_excel(writer, sheet_name=f"Parte_{i//500000 + 1}", index=False)

print(f"Archivo combinado guardado en: {ruta_salida}")
